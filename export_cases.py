import os
import re
import sys
import time
import random
import traceback
from datetime import datetime

from openpyxl import load_workbook, Workbook
from playwright.sync_api import sync_playwright


INPUT_EXCEL_NAME = "input.xlsx"
OUTPUT_EXCEL_NAME = "output.xlsx"

CDP_URL = "http://127.0.0.1:9222"

WORKBENCH_URL = (
    "https://fls-aflm-af.pingan.com.cn/web/default/"
    "assetmanagement/index.html#/workbench/phone/list"
)

CONTACT_RESULTS = [
    "无人接听",
    "占线",
    "电话拒接",
    "无法接通",
    "关机",
]


def app_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)

    return os.path.dirname(os.path.abspath(__file__))


INPUT_EXCEL = os.path.join(app_dir(), INPUT_EXCEL_NAME)
OUTPUT_EXCEL = os.path.join(app_dir(), OUTPUT_EXCEL_NAME)


def pause_exit():
    print("\n按回车键退出窗口...")
    input()


def read_excel():
    wb = load_workbook(INPUT_EXCEL)
    ws = wb.active

    headers = {}

    for idx, cell in enumerate(ws[1]):
        if cell.value:
            headers[str(cell.value).strip()] = idx

    print("当前识别到的表头：", list(headers.keys()))

    contract_keys = [
        "合同编号",
        "合同号",
        "contractNo",
        "单号",
    ]

    contract_index = next(
        (headers[k] for k in contract_keys if k in headers),
        None,
    )

    if contract_index is None:
        raise Exception("Excel 缺少合同编号列")

    tasks = []

    for row in ws.iter_rows(min_row=2):

        contract_no = row[contract_index].value

        if contract_no:
            tasks.append({
                "contract_no": str(contract_no).strip()
            })

    print(f"读取到 {len(tasks)} 条数据")

    return tasks


def save_results(results):

    wb = Workbook()
    ws = wb.active

    ws.title = "结果"

    ws.append([
        "合同编号",
        "姓名",
        "电话号码",
        "状态",
        "错误信息",
    ])

    for item in results:

        ws.append([
            item.get("contract_no", ""),
            item.get("name", ""),
            item.get("phone", ""),
            item.get("status", ""),
            item.get("error", ""),
        ])

    try:

        wb.save(OUTPUT_EXCEL)

        print(f"结果已保存：{OUTPUT_EXCEL}")

    except PermissionError:

        alt_path = OUTPUT_EXCEL.replace(
            ".xlsx",
            f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )

        wb.save(alt_path)

        print(f"output.xlsx 正在被打开，已另存为：{alt_path}")


def get_page(playwright):

    browser = playwright.chromium.connect_over_cdp(CDP_URL)

    context = browser.contexts[0]

    pages = [
        p for p in context.pages
        if not p.is_closed()
    ]

    page = pages[0] if pages else context.new_page()

    return browser, context, page


def ensure_page(playwright, page):

    if page is None or page.is_closed():

        print("页面已关闭，重新连接 Chrome...")

        _, _, page = get_page(playwright)

    return page


def click_workbench_tab(page):

    print("开始切换到电催工作台...")

    for attempt in range(3):

        try:

            success = page.evaluate(
                """
                () => {

                    const tabs = Array.from(
                        document.querySelectorAll('.ant-tabs-tab')
                    )

                    const target = tabs.find(tab => {

                        const text = (
                            tab.innerText || ''
                        ).trim()

                        return text === '电催工作台'
                    })

                    if (!target) {
                        return false
                    }

                    target.scrollIntoView({
                        block: 'center',
                        inline: 'center'
                    })

                    target.click()

                    return true
                }
                """
            )

            if success:

                print("已点击电催工作台 TAB")

                time.sleep(5)

                try:

                    page.wait_for_selector(
                        'input.ant-input[placeholder*="批量搜索"]',
                        timeout=15000,
                    )

                    print("工作台恢复成功")

                    return

                except Exception:
                    print("TAB切换后没恢复搜索框")

            print("开始使用 URL 兜底恢复工作台...")

            page.goto(
                WORKBENCH_URL,
                wait_until="domcontentloaded",
                timeout=60000,
            )

            time.sleep(8)

            page.wait_for_selector(
                'input.ant-input[placeholder*="批量搜索"]',
                timeout=45000,
            )

            print("已通过 URL 恢复工作台")

            return

        except Exception as e:

            print(f"第 {attempt + 1} 次切换工作台失败：{e}")

            if attempt == 1:

                try:

                    print("刷新页面中...")

                    page.reload(
                        wait_until="domcontentloaded",
                        timeout=60000,
                    )

                    time.sleep(8)

                except Exception:
                    pass

        time.sleep(3)

    raise Exception("切换到电催工作台失败")


def search_contract(page, contract_no):

    print(f"开始搜索合同：{contract_no}")

    page.wait_for_selector(
        'input.ant-input[placeholder*="批量搜索"]',
        timeout=30000,
    )

    contract_input = page.locator(
        'input.ant-input[placeholder*="批量搜索"]'
    ).first

    contract_input.click(force=True)

    contract_input.press("Control+A")

    contract_input.press("Backspace")

    contract_input.fill(contract_no)

    print(f"已输入合同编号：{contract_no}")

    query_btn = page.locator(
        'button.ant-btn-primary:has-text("查 询")'
    ).first

    query_btn.click(force=True)

    print("已点击查询")

    row_selector = f'tr[data-row-key="{contract_no}"]'

    try:

        page.wait_for_selector(
            row_selector,
            timeout=15000,
        )

    except Exception:

        print("未查询到合同，自动跳过")

        return False

    row = page.locator(row_selector).first

    contract_link = row.locator(
        "a",
        has_text=contract_no,
    ).first

    contract_link.scroll_into_view_if_needed()

    contract_link.click(force=True)

    print("已点击合同编号进入详情")

    time.sleep(5)

    return True


def wait_detail_ready(page):

    page.wait_for_selector(
        ".record",
        timeout=30000,
    )

    page.wait_for_selector(
        ".call-out",
        timeout=30000,
    )

    time.sleep(3)

    print("详情页数据已加载")


def get_current_status(page):

    try:

        text = page.locator(
            "div.ant-select.current-status-value"
        ).first.inner_text(timeout=5000)

        text = text.strip()

        print("当前状态：", text)

        return text

    except Exception:

        return ""


def switch_status_to_idle(page):

    current = get_current_status(page)

    if "空闲" in current:

        print("当前已经是空闲状态")

        return

    print("当前不是空闲，开始切换为空闲...")

    status_select = page.locator(
        "div.ant-select.current-status-value"
    ).first

    status_select.click(force=True)

    time.sleep(2)

    dropdown = page.locator(
        ".ant-select-dropdown:not(.ant-select-dropdown-hidden)"
    ).last

    dropdown.wait_for(
        state="visible",
        timeout=10000,
    )

    option = dropdown.locator(
        ".ant-select-item-option"
    ).filter(
        has_text="空闲"
    ).last

    option.click(force=True)

    time.sleep(3)

    print("状态已切换为空闲")


def ensure_idle_status(page):

    for i in range(3):

        try:

            current = get_current_status(page)

            if "空闲" in current:
                return

            switch_status_to_idle(page)

            current = get_current_status(page)

            if "空闲" in current:
                return

        except Exception as e:

            print(f"第 {i + 1} 次切换失败：{e}")

        time.sleep(2)

    raise Exception("无法切换为空闲状态")


def select_outbound_number(page):

    caller_select = page.locator(
        "div.ant-select.dial-caller-select"
    ).first

    caller_select.wait_for(
        state="visible",
        timeout=20000,
    )

    current_text = caller_select.inner_text().strip()

    print("当前外显号码：", current_text)

    if current_text and "请选择" not in current_text:
        return

    caller_select.click(force=True)

    time.sleep(2)

    dropdown = page.locator(
        ".ant-select-dropdown:not(.ant-select-dropdown-hidden)"
    ).last

    option = dropdown.locator(
        ".ant-select-item-option"
    ).first

    option.click(force=True)

    time.sleep(2)

    print("已选择外显号码")


def reveal_name_and_phone(page):

    try:

        icons = page.locator(
            "span.show.toggle-des"
        )

        count = icons.count()

        for i in range(count):

            try:

                icons.nth(i).click(force=True)

                time.sleep(0.5)

            except Exception:
                pass

    except Exception:
        pass


def get_name_from_page(page):

    try:

        reveal_name_and_phone(page)

        rows = page.locator(
            "tr.ant-table-row"
        )

        count = rows.count()

        for i in range(count):

            row = rows.nth(i)

            text = row.inner_text()

            if "承租人" not in text:
                continue

            spans = row.locator("span[title]")

            s_count = spans.count()

            for j in range(s_count):

                try:

                    value = spans.nth(j).get_attribute("title")

                    if value and not re.fullmatch(r"[a-f0-9]{32}", value):

                        if not re.search(r"\d", value):

                            return value

                except Exception:
                    pass

    except Exception:
        pass

    return ""


def get_real_phone(page):

    try:

        reveal_name_and_phone(page)

        rows = page.locator(
            "tr.ant-table-row"
        )

        count = rows.count()

        for i in range(count):

            row = rows.nth(i)

            text = row.inner_text()

            if "承租人" not in text:
                continue

            spans = row.locator("span[title]")

            s_count = spans.count()

            for j in range(s_count):

                try:

                    value = spans.nth(j).get_attribute("title")

                    if value and re.fullmatch(r"\d{11}", value):

                        return value

                except Exception:
                    pass

    except Exception:
        pass

    return ""


def click_call_btn(page):

    btn = page.locator(
        ".call-out img[src*='contractMakeCall']"
    ).first

    btn.scroll_into_view_if_needed()

    btn.click(force=True)

    time.sleep(2)

    try:

        call_btn = page.locator(
            "button.call-button:has-text('呼叫')"
        ).first

        if call_btn.is_visible(timeout=3000):

            call_btn.click(force=True)

            print("已点击顶部呼叫")

    except Exception:
        pass

    print("已点击拨打")


def hang_up(page):

    print("等待 3 秒后挂断...")

    time.sleep(3)

    for i in range(5):

        try:

            btn = page.locator(
                "button.call-button:has-text('挂断')"
            ).first

            if btn.is_visible(timeout=2000):

                btn.scroll_into_view_if_needed()

                btn.click(force=True)

                print("已挂断")

                time.sleep(2)

                return

        except Exception as e:

            print(f"挂断重试 {i + 1}：{e}")

        time.sleep(2)

    raise Exception("挂断失败")


def select_dropdown_value(page, label_text, target_text):

    form_item = page.locator(
        f".ant-form-item:has(label[title='{label_text}'])"
    ).last

    form_item.scroll_into_view_if_needed()

    select_root = form_item.locator(
        ".ant-select:not(.ant-select-disabled)"
    ).first

    select_root.click(force=True)

    time.sleep(1)

    dropdown = page.locator(
        ".ant-select-dropdown:not(.ant-select-dropdown-hidden)"
    ).last

    dropdown.wait_for(
        state="visible",
        timeout=10000,
    )

    options = dropdown.locator(
        ".ant-select-item-option"
    )

    count = options.count()

    found = False

    for i in range(count):

        option = options.nth(i)

        try:

            text = option.inner_text().strip()

            if text == target_text:

                option.scroll_into_view_if_needed()

                option.click(force=True)

                found = True

                break

        except Exception:
            pass

    if not found:
        raise Exception(f"{label_text} 没找到选项：{target_text}")

    time.sleep(1)

    print(f"已选择：{label_text} -> {target_text}")


def fill_collection_form(page):

    page.wait_for_selector(
        "#riskType",
        timeout=30000,
    )

    page.wait_for_selector(
        "#contactResult",
        timeout=30000,
    )

    try:

        select_dropdown_value(
            page,
            "风险分类",
            "失联",
        )

    except Exception as e:

        print("风险分类选择失败：", e)

    result = random.choice(CONTACT_RESULTS)

    try:

        select_dropdown_value(
            page,
            "联络结果",
            result,
        )

    except Exception as e:

        print("联络结果选择失败：", e)


def submit_form(page):

    btn = page.locator(
        "button.ant-btn-primary:has-text('提 交')"
    ).last

    btn.scroll_into_view_if_needed()

    btn.click(force=True)

    time.sleep(3)

    print("已提交")


def close_detail_tab(page):

    try:

        page.evaluate(
            """
            () => {

                const tabs = Array.from(
                    document.querySelectorAll('.ant-tabs-tab')
                )

                const detailTab = tabs.find(tab => {

                    const text = (
                        tab.innerText || ''
                    ).trim()

                    return text.includes('详情')
                })

                if (!detailTab) {
                    return
                }

                const closeBtn = detailTab.querySelector(
                    '.ant-tabs-tab-remove'
                )

                if (closeBtn) {
                    closeBtn.click()
                }
            }
            """
        )

        print("已关闭详情 TAB")

        time.sleep(2)

    except Exception as e:

        print("关闭详情 TAB 失败：", e)


def process_case(page, task):

    contract_no = task["contract_no"]

    print(f"开始处理：{contract_no}")

    click_workbench_tab(page)

    found = search_contract(page, contract_no)

    if not found:

        return {
            "contract_no": contract_no,
            "name": "",
            "phone": "",
            "status": "未找到合同",
            "error": "",
        }

    wait_detail_ready(page)

    ensure_idle_status(page)

    select_outbound_number(page)

    ensure_idle_status(page)

    name = get_name_from_page(page)

    phone = get_real_phone(page)

    print("姓名：", name)

    print("手机号：", phone)

    click_call_btn(page)

    hang_up(page)

    fill_collection_form(page)

    submit_form(page)

    close_detail_tab(page)

    print(f"完成：{contract_no}")

    return {
        "contract_no": contract_no,
        "name": name,
        "phone": phone,
        "status": "成功",
        "error": "",
    }


def main():

    try:

        if not os.path.exists(INPUT_EXCEL):

            print(f"没找到 input.xlsx：{INPUT_EXCEL}")

            return

        tasks = read_excel()

        results = []

        with sync_playwright() as p:

            browser, context, page = get_page(p)

            print("=" * 50)
            print("请确保：")
            print("1.Chrome 已启动远程调试")
            print("2.已经登录系统")
            print("3.当前可访问工作台")
            print("=" * 50)

            input("确认后按回车继续...")

            for index, task in enumerate(tasks, start=1):

                print(f"\n[{index}/{len(tasks)}]")

                try:

                    page = ensure_page(p, page)

                    result = process_case(
                        page,
                        task,
                    )

                except Exception as e:

                    print("处理失败")

                    print(traceback.format_exc())

                    result = {
                        "contract_no": task.get("contract_no", ""),
                        "name": "",
                        "phone": "",
                        "status": "失败",
                        "error": str(e),
                    }

                results.append(result)

                save_results(results)

                time.sleep(2)

            print("\n全部完成")

            print(f"结果文件：{OUTPUT_EXCEL}")

    except Exception:

        print("程序发生未捕获异常：")

        print(traceback.format_exc())

    finally:

        pause_exit()


if __name__ == "__main__":
    main()
