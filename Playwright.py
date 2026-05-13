import os
import time
import traceback
from urllib.parse import urlparse, parse_qs, unquote

from openpyxl import load_workbook, Workbook
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError


INPUT_EXCEL = "input.xlsx"
OUTPUT_EXCEL = "output.xlsx"

CHROME_PATHS = [
    r"C:\Program Files\Google\Chrome\Application\chrome.exe",
    r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
]

USER_DATA_DIR = "./chrome-user-data"


def find_chrome():
    for path in CHROME_PATHS:
        if os.path.exists(path):
            return path
    raise FileNotFoundError("没找到 Chrome，请确认电脑已安装 Google Chrome")


def get_cell(row, headers, *names):
    for name in names:
        if name in headers:
            return row[headers[name]].value
    return None


def parse_contract_no(url):
    try:
        query = parse_qs(urlparse(url).query)
        return query.get("contractNo", [""])[0]
    except Exception:
        return ""


def parse_name_from_url(url):
    try:
        query = parse_qs(urlparse(url).query)
        return unquote(query.get("loanName", [""])[0])
    except Exception:
        return ""


def read_tasks():
    wb = load_workbook(INPUT_EXCEL)
    ws = wb.active

    headers = {}
    for idx, cell in enumerate(ws[1]):
        if cell.value:
            headers[str(cell.value).strip()] = idx

    tasks = []

    for row in ws.iter_rows(min_row=2):
        detail_url = get_cell(row, headers, "详情页URL", "url", "URL", "detailUrl")
        contract_no = get_cell(row, headers, "合同编号", "contractNo", "单号")

        if not detail_url:
            continue

        detail_url = str(detail_url).strip()
        contract_no = str(contract_no).strip() if contract_no else parse_contract_no(detail_url)

        tasks.append({
            "contract_no": contract_no,
            "url": detail_url,
            "name": parse_name_from_url(detail_url),
        })

    return tasks


def select_ant_option(page, input_id, option_text):
    page.locator(f"input#{input_id}").locator("xpath=ancestor::div[contains(@class,'ant-select')]").click()
    page.locator(
        ".ant-select-dropdown:not(.ant-select-dropdown-hidden) .ant-select-item-option",
        has_text=option_text
    ).last.click()


def select_ant_any_option(page, input_id):
    page.locator(f"input#{input_id}").locator("xpath=ancestor::div[contains(@class,'ant-select')]").click()
    page.locator(
        ".ant-select-dropdown:not(.ant-select-dropdown-hidden) .ant-select-item-option"
    ).first.click()


def reveal_and_get_phone(page):
    try:
        page.locator("span.show.toggle-des").first.click(timeout=3000)
        time.sleep(1)
    except Exception:
        pass

    phone_text = ""

    try:
        phone_text = page.locator(".call-out").first.inner_text(timeout=3000).strip()
    except Exception:
        pass

    return phone_text.replace("\n", " ").strip()


def get_name_from_page(page, fallback_name):
    try:
        text = page.locator("text=承租人").first.locator("xpath=ancestor::*[contains(@class,'ant-card') or contains(@class,'basic') or contains(@class,'ant-tabs')]").inner_text(timeout=3000)
        return fallback_name
    except Exception:
        return fallback_name


def process_one(page, task):
    contract_no = task["contract_no"]
    url = task["url"]

    page.goto(url, wait_until="domcontentloaded")
    page.wait_for_load_state("networkidle", timeout=20000)
    time.sleep(2)

    # 先点眼睛，获取真实电话
    phone = reveal_and_get_phone(page)

    # 点击电话呼出按钮
    call_btn = page.locator(".call-out img[src*='contractMakeCall']").first
    call_btn.click(timeout=10000)

    # 外显号码，如果默认已有济南，可以不动；这里尝试选择济南
    try:
        select_ant_option(page, "rc_select_3", "济南")
    except Exception:
        pass

    # 等三秒
    time.sleep(3)

    # 点击挂断
    try:
        hangup = page.locator("button.call-button:has-text('挂断')").first
        hangup.click(force=True, timeout=5000)
    except Exception:
        page.locator("button.call-button").first.click(force=True, timeout=5000)

    time.sleep(1)

    # 风险分类：失联
    select_ant_option(page, "riskType", "失联")

    # 联络结果：随便选第一个
    select_ant_any_option(page, "contactResult")

    # 提交
    page.locator("button.ant-btn-primary:has-text('提 交')").first.click(timeout=10000)

    time.sleep(2)

    return {
        "contract_no": contract_no,
        "name": task["name"],
        "phone": phone,
        "status": "成功",
        "error": "",
    }


def save_results(results):
    wb = Workbook()
    ws = wb.active
    ws.title = "结果"

    ws.append(["合同编号", "姓名", "电话", "状态", "错误信息"])

    for item in results:
        ws.append([
            item.get("contract_no", ""),
            item.get("name", ""),
            item.get("phone", ""),
            item.get("status", ""),
            item.get("error", ""),
        ])

    wb.save(OUTPUT_EXCEL)


def main():
    tasks = read_tasks()
    results = []

    chrome_path = find_chrome()

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=USER_DATA_DIR,
            executable_path=chrome_path,
            headless=False,
            args=["--start-maximized"],
            viewport=None,
        )

        page = context.pages[0] if context.pages else context.new_page()

        print("第一次运行如果没登录，请先手动登录。登录完成后回车继续。")
        input()

        for index, task in enumerate(tasks, start=1):
            print(f"开始处理 {index}/{len(tasks)}：{task['contract_no']}")

            try:
                result = process_one(page, task)
            except Exception as e:
                result = {
                    "contract_no": task.get("contract_no", ""),
                    "name": task.get("name", ""),
                    "phone": "",
                    "status": "失败",
                    "error": str(e),
                }
                print(traceback.format_exc())

            results.append(result)
            save_results(results)

        context.close()

    print(f"全部完成，结果已保存到 {OUTPUT_EXCEL}")


if __name__ == "__main__":
    main()
