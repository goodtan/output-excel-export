import base64
import hashlib
import time
from datetime import datetime
import requests
from Crypto.Cipher import AES
from Crypto.Util.Padding import unpad
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


TOKEN = "370c19a134924cef4c6c6f03cfb0707c"

CRYPT_IV = "+noB4IVKOqQA7T2iH9MGAw=="
CRYPT_KEY = "563BCDCFACFEF5C74676C7EBFBBDB613"

BASE_URL = "https://server.xingeguanli.com/osapi/Cases/index"


COLUMNS = [
    ("案件ID", "id"),
    ("案件状态", "case_status_text"),
    ("批次号【委托方】", "bill_no_text"),
    ("失信被执行人", "is_shixin_text"),
    ("客户姓名", "case_name"),
    ("合同号", "contract"),
    ("身份证【户籍地】", "idcard_text"),
    ("手机号", "case_phone"),
    ("关联公众号", "wxamp"),
    ("剩余应还", "amount_payable"),

    ("账期", "period"),
    ("逾期天数", "overdue_days"),
    ("催收员", "member_name"),
    ("地区", "case_area"),
    ("跟进状态", "follow_status_text"),
    ("催收状态", "coll_status_text"),
    ("委案日期", "entrust_start_text"),
    ("退案日期", "entrust_end_text"),
    ("分配时间", "follow_time_text"),
    ("未跟进天数", "unfollow_days"),
    ("最后跟进时间", "last_follow_time_text"),
]


def decrypt_data(encrypted_data: str):
    final_key = "MC.1888@#!1" + CRYPT_KEY

    key = hashlib.sha256(final_key.encode("utf-8")).digest()
    iv = base64.b64decode(CRYPT_IV)
    cipher_text = base64.b64decode(encrypted_data)

    cipher = AES.new(key, AES.MODE_CBC, iv)
    decrypted = unpad(cipher.decrypt(cipher_text), AES.block_size)

    # 解密出来还是一层 Base64
    base64_text = decrypted.decode("utf-8")
    json_text = base64.b64decode(base64_text).decode("utf-8")

    import json
    return json.loads(json_text)


def format_time(value):
    if not value or value == 0 or value == "--":
        return ""
    return datetime.fromtimestamp(int(value)).strftime("%Y-%m-%d %H:%M:%S")


def map_case_status(value):
    mapping = {
        1: "正常",
        2: "暂停",
        3: "退案",
    }
    return mapping.get(value, value or "")


def map_coll_status(value):
    mapping = {
        20: "待跟进",
    }
    return mapping.get(value, value or "")


def map_shixin(value):
    mapping = {
        1: "是",
        2: "否",
    }
    return mapping.get(value, "")


def transform_row(item):
    bill_no = item.get("bill_no") or ""
    mem_primary_name = item.get("mem_primary_name") or ""

    case_idcard = item.get("case_idcard_asterisk") or item.get("case_idcard") or ""
    idcard_area = item.get("idcard_area") or ""

    return {
        "id": item.get("id", ""),
        "case_status_text": map_case_status(item.get("case_status")),
        "bill_no_text": f"{bill_no}【{mem_primary_name}】" if mem_primary_name else bill_no,
        "is_shixin_text": map_shixin(item.get("is_shixin")),
        "case_name": item.get("case_name", ""),
        "contract": item.get("contract", ""),
        "idcard_text": f"{case_idcard}【{idcard_area}】" if idcard_area else case_idcard,
        "case_phone": item.get("case_phone", ""),
        "wxamp": item.get("wxamp", ""),
        "amount_payable": item.get("amount_payable") or item.get("new_entrust_money") or "",

        "period": item.get("period", ""),
        "overdue_days": item.get("overdue_days", ""),
        "member_name": item.get("member_name", ""),
        "case_area": item.get("case_area", ""),
        "follow_status_text": "已跟进" if item.get("follow_time") else "未跟进",
        "coll_status_text": map_coll_status(item.get("coll_status")),
        "entrust_start_text": format_time(item.get("entrust_start")),
        "entrust_end_text": format_time(item.get("entrust_end")),
        "follow_time_text": format_time(item.get("follow_time")),
        "unfollow_days": item.get("unfollow_days", ""),
        "last_follow_time_text": format_time(item.get("note_time") or item.get("follow_time")),
    }


def fetch_page(page, per_page=10):
    headers = {
        "accept": "*/*",
        "origin": "https://os.xingeguanli.com",
        "referer": "https://os.xingeguanli.com/",
        "token": TOKEN,
        "x-requested-with": "XMLHttpRequest",
        "user-agent": "Mozilla/5.0",
    }

    params = {
        "page": page,
        "perPage": per_page,
        "status": 0,
        "id": "",
        "sortstatus": "",
        "sorttype": "",
        "snatch_id": "",
        "case_name": "",
        "case_phone": "",
        "case_idcard": "",
        "contract": "",
        "bill_nos": "",
        "entrust_id": "",
        "case_status": "",
        "coll_status": "",
        "case_area": "",
        "period": "",
        "member_ids": "",
        "new_entrust_money": "0,0",
        "entrust_money": "0,0",
        "unfollow_days": "0,0",
        "follow_time": "",
        "entrust_start": "",
        "entrust_end": "",
        "phone_clean": "",
        "is_settle": "",
        "score1": "",
        "score2": "",
        "risk_tag": "",
        "case_user_ask": "",
        "tag_kelian": "",
        "tag_case": "",
        "is_shixin": "",
        "idcard_area": "",
    }

    response = requests.get(
        BASE_URL,
        headers=headers,
        params=params,
        timeout=180,
    )

    response.raise_for_status()

    body = response.json()

    if body.get("code") != 1:
        raise Exception(body.get("msg", "接口请求失败"))

    return decrypt_data(body["data"])


def export_excel(data_list):
    wb = Workbook()
    ws = wb.active
    ws.title = "案件列表"

    headers = [item[0] for item in COLUMNS]
    keys = [item[1] for item in COLUMNS]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for item in data_list:
        row = transform_row(item)
        ws.append([row.get(key, "") for key in keys])

    for col_index, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_index)
        ws.column_dimensions[col_letter].width = max(len(header) * 2, 16)

    filename = f"案件列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)

    print(f"Excel 导出完成：{filename}")


def main():
    all_data = []
    per_page = 10

    page = 1

    while True:
        print(f"正在请求第 {page} 页...")

        decrypted = fetch_page(page, per_page)

        data_list = decrypted.get("data", [])
        count = int(decrypted.get("count", 0))

        all_data.extend(data_list)

        print(f"第 {page} 页 {len(data_list)} 条，累计 {len(all_data)} 条，总数 {count}")

        if not data_list or len(all_data) >= count:
            break

        page += 1
        time.sleep(0.5)

    export_excel(all_data)


if __name__ == "__main__":
    main()
